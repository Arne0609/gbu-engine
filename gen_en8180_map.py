# -*- coding: utf-8 -*-
"""Erzeugt die EN-81-80-Sicht auf den mehrfragigen Katalog:
  python3 gen_en8180_map.py
Eingaben : en8180_content.py, norm_81_20_mf.json
Ausgaben : en8180_map.json                                   (Referenz/Engine)
           ../gbu_aufzug_app/lib/data/en8180_katalog.dart    (App, erzeugt)
           GBU_EN8180_Zuordnung.xlsx                         (Gegenlesung)

Prüft dabei, dass jede zugeordnete MF-Gefährdung im Katalog existiert, und
listet die Gefährdungssituationen ohne Entsprechung als Lücke auf.
"""
import json, os, sys
from en8180_content import (ZUORDNUNG, ZEITPLAN, prioritaeten,
                            GEGENGELESEN, GEGENGELESEN_HINWEIS)

HERE = os.path.dirname(os.path.abspath(__file__))
mf = json.load(open(os.path.join(HERE, 'norm_81_20_mf.json'), encoding='utf-8'))
H = {h['code']: h for h in mf['hazards']}
PRIO = prioritaeten()

fehler = []
for nr, (abschnitt, titel, hazards, deckung, bem) in ZUORDNUNG.items():
    for c in hazards:
        if c not in H:
            fehler.append('Nr. %d: MF-Gefährdung %s unbekannt' % (nr, c))
    if deckung not in ('voll', 'teilweise', 'offen'):
        fehler.append('Nr. %d: unbekannte Deckung %r' % (nr, deckung))
    if deckung == 'offen' and hazards:
        fehler.append('Nr. %d: „offen" mit Zuordnung' % nr)
    if deckung != 'offen' and not hazards:
        fehler.append('Nr. %d: ohne Zuordnung, aber nicht „offen"' % nr)
if fehler:
    print('\n'.join(fehler))
    sys.exit(1)

punkte = []
for nr in sorted(ZUORDNUNG):
    abschnitt, titel, hazards, deckung, bem = ZUORDNUNG[nr]
    prio, felder = PRIO[nr]
    punkte.append({
        'nr': nr,
        'abschnitt': abschnitt,
        'titel': titel,
        'prioritaet': prio,
        'risikoprofil': felder,
        'zeitplan': ZEITPLAN[prio],
        'hazards': hazards,
        'deckung': deckung,
        'bemerkung': bem,
    })

daten = {
    'quelle': 'DIN EN 81-80:2004-02, Tabelle 1 sowie Anhang A '
              '(Tabellen A.1 und A.2)',
    'katalog': mf['rule_version'],
    'gegengelesen': GEGENGELESEN,
    'gegengelesen_hinweis': GEGENGELESEN_HINWEIS,
    'punkte': punkte,
}
ziel = os.path.join(HERE, 'en8180_map.json')
json.dump(daten, open(ziel, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

# ---- Dart-Katalog (erzeugt) ------------------------------------------------
def dart_str(s):
    return "'" + s.replace('\\', '\\\\').replace("'", "\\'").replace('$', r'\$') + "'"


zeilen = [
    '// AUTOMATISCH ERZEUGT von engine_model/gen_en8180_map.py – nicht von Hand',
    '// ändern. Quelle: %s.' % daten['quelle'],
    '// Nummern, Abschnitte und Prioritätsstufen stammen aus der Norm; die',
    '// Kurzbezeichnungen sind eigene Formulierungen (kein Normtext).',
    '',
    "import 'en8180_map.dart';",
    '',
    '/// Die 74 Gefährdungssituationen der DIN EN 81-80 mit ihrer Zuordnung zum',
    '/// mehrfragigen Katalog (%s).' % mf['rule_version'],
    "const String en8180Gegengelesen = '%s';" % GEGENGELESEN,
    '',
    'const List<En8180Punkt> en8180Punkte = [',
]
for p in punkte:
    zeilen.append('  En8180Punkt(')
    zeilen.append('    nr: %d,' % p['nr'])
    zeilen.append('    abschnitt: %s,' % dart_str(p['abschnitt']))
    zeilen.append('    titel: %s,' % dart_str(p['titel']))
    zeilen.append('    prioritaet: %s,' % dart_str(p['prioritaet']))
    zeilen.append('    zeitplan: %s,' % dart_str(p['zeitplan']))
    zeilen.append('    deckung: %s,' % dart_str(p['deckung']))
    if p['hazards']:
        zeilen.append('    hazards: [%s],' %
                      ', '.join(dart_str(c) for c in p['hazards']))
    if p['bemerkung']:
        zeilen.append('    bemerkung: %s,' % dart_str(p['bemerkung']))
    zeilen.append('  ),')
zeilen.append('];')
zeilen.append('')

dart_ziel = os.path.join(HERE, '..', 'gbu_aufzug_app', 'lib', 'data',
                         'en8180_katalog.dart')
if len(sys.argv) > 1:
    dart_ziel = sys.argv[1]
open(dart_ziel, 'w', encoding='utf-8').write('\n'.join(zeilen))

# ---- Excel zur Gegenlesung -------------------------------------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

if Workbook is not None:
    FONT = 'Arial'
    HEAD = PatternFill('solid', fgColor='1F3A5F')
    INPUT = PatternFill('solid', fgColor='FFF2CC')
    LUECKE = PatternFill('solid', fgColor='FCE4E4')
    STRIPE = PatternFill('solid', fgColor='F3F5F8')
    thin = Side(style='thin', color='C9D1DB')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    WRAP = Alignment(wrap_text=True, vertical='top')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Zuordnung'
    cols = ['Nr.', 'Abschnitt EN 81-80', 'Gefährdungssituation (Kurzbezeichnung)',
            'Priorität', 'Risikoprofil A.1', 'Zeitplan A.2',
            'MF-Gefährdung(en)', 'Titel der MF-Gefährdung(en)', 'Deckung',
            'Bemerkung', 'Zuordnung OK?', 'Korrektur']
    widths = [5, 14, 52, 10, 14, 34, 20, 52, 11, 40, 12, 34]
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        cell.fill = HEAD
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30
    for r, p in enumerate(punkte, 2):
        werte = [p['nr'], p['abschnitt'], p['titel'],
                 p['prioritaet'] or 'nicht eingestuft',
                 ', '.join(p['risikoprofil']) or '-', p['zeitplan'],
                 ', '.join(p['hazards']) or '-',
                 '\n'.join(H[c]['title'] for c in p['hazards']) or '-',
                 p['deckung'], p['bemerkung'], None, None]
        for c, v in enumerate(werte, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = WRAP
            cell.border = BORDER
            if c in (11, 12):
                cell.fill = INPUT
            elif p['deckung'] != 'voll':
                cell.fill = LUECKE
            elif r % 2 == 0:
                cell.fill = STRIPE
    dv = DataValidation(type='list', formula1='"OK,Ändern"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add('K2:K%d' % (len(punkte) + 1))
    ws.auto_filter.ref = 'A1:L%d' % (len(punkte) + 1)

    ws2 = wb.create_sheet('Lesehinweise')
    ws2.column_dimensions['A'].width = 120
    luecken = [p for p in punkte if p['deckung'] == 'offen']
    teil = [p for p in punkte if p['deckung'] == 'teilweise']
    zeilen2 = [
        ('EN 81-80 als Sicht auf den mehrfragigen Katalog (%s)' % mf['rule_version'], True),
        ('Quelle: %s' % daten['quelle'], False),
        ('', False),
        ('Der eingefrorene GBU-Typ „vereinfacht (EN 81-80)" wird nicht durch einen zweiten '
         'Fragebogen ersetzt: Eine Bestandsanlage wird einmal nach EN 81-20 (mehrfragig) '
         'erhoben, der Bericht weist den Nachrüstbedarf nach EN 81-80 zusätzlich mit '
         'Nummer, Abschnitt und Priorität der Norm aus.', False),
        ('', False),
        ('%d der 74 Gefährdungssituationen sind vollständig abgedeckt, %d teilweise, '
         '%d gar nicht (Spalten rot hinterlegt).'
         % (len(punkte) - len(luecken) - len(teil), len(teil), len(luecken)), False),
        ('Lücken im MF-Katalog: %s' % ', '.join(
            'Nr. %d %s' % (p['nr'], p['titel']) for p in luecken), False),
        ('Teilweise: %s' % ', '.join(
            'Nr. %d %s' % (p['nr'], p['titel']) for p in teil), False),
        ('', False),
        ('Priorität nach Tabelle A.2 aus dem Risikoprofil A.1 abgeleitet (Schwere/Häufigkeit); '
         'erscheint eine Nummer dort mehrfach, gilt die höhere Stufe. Die Nummern 2, 4 und 5 '
         'sind in A.1 nicht enthalten (besondere Anforderungen nach 5.1.5) und deshalb nicht '
         'eingestuft.', False),
        ('', False),
        ('Gelbe Spalten ausfüllen: „Zuordnung OK?" = OK/Ändern, Korrektur frei.', False),
        ('', False),
        ('Stand der Gegenlesung: %s – %s' % (GEGENGELESEN, GEGENGELESEN_HINWEIS), True),
    ]
    for i, (t, b) in enumerate(zeilen2, 1):
        cell = ws2.cell(row=i, column=1, value=t)
        cell.font = Font(name=FONT, bold=b, size=11 if b else 10)
        cell.alignment = WRAP
    wb.save(os.path.join(HERE, 'GBU_EN8180_Zuordnung.xlsx'))

from collections import Counter
print('geschrieben: en8180_map.json, %s, GBU_EN8180_Zuordnung.xlsx'
      % os.path.normpath(dart_ziel))
print('Punkte: %d | Deckung: %s | Priorität: %s'
      % (len(punkte), dict(Counter(p['deckung'] for p in punkte)),
         dict(Counter(p['prioritaet'] or 'nicht eingestuft' for p in punkte))))
print('abgedeckte MF-Gefährdungen: %d von %d'
      % (len({c for p in punkte for c in p['hazards']}), len(H)))
