# -*- coding: utf-8 -*-
"""Übernimmt die zurückgegebene Regelprüfung des Cyber-Typs:
  python3 apply_cy_regelpruefung.py [GBU_Cyber_Regelpruefung.xlsx] [--datum JJJJ-MM-TT]
Liest Blatt „Muster" (M1…M6, gilt für alle Komponenten CY-C01…C10) und Blatt
„Regeln" (Einzelentscheidung geht vor) und schreibt cy_content/regelfreigabe.py.
Danach: python3 gen_cy_catalog.py  (setzt freigegebene Regeln auf VERIFIED).
"""
import json, os, re, sys, datetime
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
argv = sys.argv[1:]
datum = datetime.date.today().isoformat()
if '--datum' in argv:
    i = argv.index('--datum')
    datum = argv[i + 1]
    # Option UND ihren Wert entfernen, sonst wird der Wert als Dateiname gelesen.
    del argv[i:i + 2]
args = [a for a in argv if not a.startswith('--')]
xlsx = args[0] if args else os.path.join(HERE, 'GBU_Cyber_Regelpruefung.xlsx')

seed = json.load(open(os.path.join(HERE, 'norm_cyber_mf.json'), encoding='utf-8'))
review = [r for r in seed['rules'] if r.get('quality_status') == 'REVIEW_REQUIRED']

def pattern_key(r):
    return re.sub(r'qc_[a-z]+_', 'qc_X_', json.dumps(r['condition'], sort_keys=True)) + '|' + r['result']

# Muster-Zuordnung wie in gen_cy_regelpruefung_xlsx.py (Reihenfolge des ersten Auftretens)
patterns = {}
for r in review:
    if r['hazard'].startswith('CY-C') and int(r['hazard'][4:]) <= 10:
        patterns.setdefault(pattern_key(r), []).append(r)
pattern_id = {k: 'M%d' % (i + 1) for i, k in enumerate(patterns)}

wb = load_workbook(xlsx, data_only=True)
VALID = {'Freigeben', 'Ändern', 'Streichen'}

def norm(v):
    v = (v or '').strip() if isinstance(v, str) else v
    return v if v in VALID else None

freigabe = {}
# 1) Muster
ws = wb['Muster']
muster_entscheidung = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    mid, ent, korr = row[0], norm(row[10]), row[11]
    if ent:
        muster_entscheidung[mid] = (ent, (korr or '').strip() if isinstance(korr, str) else '')
for k, rs in patterns.items():
    if pattern_id[k] in muster_entscheidung:
        for r in rs:
            freigabe[r['code']] = muster_entscheidung[pattern_id[k]]
# 2) Einzelregeln (gehen vor)
ws = wb['Regeln']
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[4] or not isinstance(row[4], str) or not row[4].startswith('CY-'):
        continue
    ent, korr = norm(row[14]), row[15]
    if ent:
        freigabe[row[4]] = (ent, (korr or '').strip() if isinstance(korr, str) else '')

known = {r['code'] for r in review}
unknown = sorted(set(freigabe) - known)
if unknown:
    print('WARNUNG: Entscheidungen für unbekannte/bereits verifizierte Regeln ignoriert:', unknown)
    for u in unknown: freigabe.pop(u)

out = os.path.join(HERE, 'cy_content', 'regelfreigabe.py')
src = open(out, encoding='utf-8').read()
head = src.split('DATUM = ')[0]
body = ['DATUM = %r' % datum, '', 'FREIGABE = {']
for code in sorted(freigabe, key=lambda c: [int(x) if x.isdigit() else x for x in re.split(r'(\d+)', c)]):
    ent, korr = freigabe[code]
    body.append('    %r: (%r, %r),' % (code, ent, korr))
body.append('}')
open(out, 'w', encoding='utf-8').write(head + '\n'.join(body) + '\n')

from collections import Counter
c = Counter(v[0] for v in freigabe.values())
print('geschrieben:', out, '| Regeln entschieden: %d von %d' % (len(freigabe), len(review)),
      '| Freigeben %d · Ändern %d · Streichen %d' % (c['Freigeben'], c['Ändern'], c['Streichen']),
      '| offen: %d' % (len(review) - len(freigabe)))
if c['Ändern'] or c['Streichen']:
    print('Hinweis: Ändern/Streichen im Inhalt (cy_content/*.py) nachziehen, dann gen_cy_catalog.py.')
    for code, (ent, korr) in sorted(freigabe.items()):
        if ent != 'Freigeben': print('  %-12s %-9s %s' % (code, ent, korr))
