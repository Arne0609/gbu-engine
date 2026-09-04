# -*- coding: utf-8 -*-
"""Erzeugt die Klärungsliste + Regelübersicht des Cyber-Typs als Excel
für die fachliche Gegenlesung:  python3 gen_cy_review_xlsx.py
Eingaben: norm_cyber_mf.json, cy_klaerung.json  ->  GBU_Cyber_Klaerungsliste.xlsx
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
seed = json.load(open(os.path.join(HERE, 'norm_cyber_mf.json'), encoding='utf-8'))
kl = json.load(open(os.path.join(HERE, 'cy_klaerung.json'), encoding='utf-8'))
OUT = os.path.join(HERE, 'GBU_Cyber_Klaerungsliste.xlsx')

Q = {q['code']: q for q in seed['questions']}
H = {h['code']: h for h in seed['hazards']}
M = {m['code']: m for m in seed['measures']}
LAB = {'HIGH': 'Hoch', 'MEDIUM': 'Mittel', 'LOW': 'Niedrig', 'NO_RISK': 'Kein Risiko',
       'NOT_APPLICABLE': 'Nicht zutreffend', 'INCOMPLETE': 'Unvollständig'}
ROLE = {'APPLICABILITY': 'Filter', 'TRIGGER': 'Auslöser', 'COMPENSATION': 'Kompensation',
        'MODIFIER': 'Modifier', 'OPTIONAL': 'optional', 'DOCUMENTATION': 'Dokumentation'}

FONT = 'Arial'
def f(bold=False, color=None, size=10):
    return Font(name=FONT, bold=bold, color=color, size=size)
HEAD_FILL = PatternFill('solid', fgColor='1F3A5F')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')
STRIPE = PatternFill('solid', fgColor='F3F5F8')
thin = Side(style='thin', color='C9D1DB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top')


def val_label(qc, v):
    q = Q.get(qc)
    if q and q.get('type') == 'SELECT':
        for o in q.get('options', []):
            if o['value'] == v:
                return '„%s"' % o['label']
    if v is True: return 'Ja'
    if v is False: return 'Nein'
    return str(v)


def qname(qc):
    q = Q.get(qc)
    if not q: return qc
    return '%s %s' % (q.get('ui_number', ''), q['text'])


def expr(e):
    if 'all' in e: return '(' + ' UND '.join(expr(x) for x in e['all']) + ')'
    if 'any' in e: return '(' + ' ODER '.join(expr(x) for x in e['any']) + ')'
    if 'not' in e: return 'NICHT ' + expr(e['not'])
    op = e['operator']
    if op in ('ANSWERED', 'NOT_ANSWERED'):
        return qname(e['question']) + (' beantwortet' if op == 'ANSWERED' else ' unbeantwortet')
    v = e.get('value')
    if op in ('IN', 'NOT_IN'):
        return '%s %s [%s]' % (qname(e['question']), 'in' if op == 'IN' else 'nicht in',
                               ', '.join(val_label(e['question'], x) for x in v))
    sym = {'EQ': '=', 'NEQ': '≠', 'GT': '>', 'GTE': '≥', 'LT': '<', 'LTE': '≤'}[op]
    return '%s %s %s' % (qname(e['question']), sym, val_label(e['question'], v))


def header(ws, cols, widths):
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30


def write_rows(ws, rows, input_cols=()):
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = f()
            cell.alignment = WRAP
            cell.border = BORDER
            if c in input_cols:
                cell.fill = INPUT_FILL
            elif r % 2 == 0:
                cell.fill = STRIPE


wb = Workbook()

# ---- Blatt 1: Klärungen ----------------------------------------------------
ws = wb.active
ws.title = 'Klaerungen'
cols = ['ID', 'Bereich', 'Thema', 'Frage', 'Vorschlag', 'Alternative', 'Grund',
        'Betroffene Gefährdungen', 'Entscheidung', 'Bemerkung / Festlegung']
header(ws, cols, [9, 14, 24, 60, 30, 34, 36, 24, 18, 40])
rows = []
for k in kl:
    rows.append([k['id'], k['bereich'], k['thema'], k['frage'], k['vorschlag'], k['alternativen'],
                 k['grund'], ', '.join(k['hazards']),
                 {'Anders': 'Anders (siehe Bemerkung)'}.get(k.get('entscheidung'), k.get('entscheidung')),
                 (k.get('festlegung') + ' (Stand ' + k.get('datum', '') + ')') if k.get('festlegung') else None])
write_rows(ws, rows, input_cols=(9, 10))
dv = DataValidation(type='list', formula1='"Vorschlag,Alternative,Anders (siehe Bemerkung)"',
                    allow_blank=True)
ws.add_data_validation(dv)
dv.add('I2:I%d' % (len(rows) + 1))
n = len(rows) + 1
r0 = n + 2
ws.cell(row=r0, column=1, value='Stand der Gegenlesung').font = f(bold=True)
ws.cell(row=r0 + 1, column=1, value='Klärungen gesamt').font = f()
ws.cell(row=r0 + 1, column=2, value='=COUNTA(A2:A%d)' % n).font = f()
ws.cell(row=r0 + 2, column=1, value='entschieden').font = f()
ws.cell(row=r0 + 2, column=2, value='=COUNTA(I2:I%d)' % n).font = f()
ws.cell(row=r0 + 3, column=1, value='offen').font = f()
ws.cell(row=r0 + 3, column=2, value='=B%d-B%d' % (r0 + 1, r0 + 2)).font = f()
ws.cell(row=r0 + 5, column=1, value='Legende: gelbe Zellen ausfüllen (Entscheidung als Auswahl, '
        'Bemerkung frei). „Vorschlag" = so wie im Seed umgesetzt; „Alternative" = die genannte '
        'Alternative; „Anders" = Festlegung in der Bemerkung.').font = f(size=9)

# ---- Blatt 2: Regeln -------------------------------------------------------
ws = wb.create_sheet('Regeln')
cols = ['Gefährdung', 'Titel', 'Baugruppe', 'Regel', 'Prio', 'Bedingung', 'Stufe', 'Evidenz',
        'Klärung', 'Sofortmaßnahme', 'Mittelfristige Maßnahme', 'Hinweis', 'Prüfung OK?', 'Korrektur']
header(ws, cols, [11, 38, 22, 13, 6, 70, 13, 16, 12, 40, 40, 30, 11, 36])
rows = []
for r in seed['rules']:
    h = H[r['hazard']]
    sofort = mittel = ''
    for mb in r.get('measures', []):
        m = M.get(mb['measure'])
        if not m: continue
        if mb.get('group_id') == 'sofort': sofort = m['title']
        elif mb.get('group_id') == 'mittel': mittel = m['title']
    notes = r.get('notes', '')
    kids = ', '.join(x for x in (notes.split('KLÄREN: ')[1].split(', ') if 'KLÄREN: ' in notes else []))
    hint = notes.split(' KLÄREN: ')[0] if not notes.startswith('KLÄREN: ') else ''
    rows.append([r['hazard'], h['title'], h.get('category', ''), r['code'], r['priority'],
                 expr(r['condition']), LAB[r['result']], r.get('evidence', ''), kids,
                 sofort, mittel, hint, None, None])
write_rows(ws, rows, input_cols=(13, 14))
dv2 = DataValidation(type='list', formula1='"OK,Ändern"', allow_blank=True)
ws.add_data_validation(dv2)
dv2.add('M2:M%d' % (len(rows) + 1))
ws.auto_filter.ref = 'A1:N%d' % (len(rows) + 1)

# ---- Blatt 3: Gefährdungen -------------------------------------------------
ws = wb.create_sheet('Gefaehrdungen')
cols = ['Code', 'Titel', 'Baugruppe', 'Erhebungsbereich', 'Aggregation', 'Gefährdungsfaktor',
        'Personengruppen', 'Fragen (Rolle · Pflicht)', 'Norm-/Quellenbezug', 'Klärung', 'Regeln']
header(ws, cols, [10, 44, 24, 28, 11, 40, 26, 70, 40, 12, 8])
rows = []
rules_by = {}
for r in seed['rules']: rules_by.setdefault(r['hazard'], []).append(r)
for h in seed['hazards']:
    qs = []
    for hq in h.get('questions', []):
        q = Q.get(hq['question'])
        req = hq.get('required_mode', 'NEVER')
        qs.append('%s [%s%s]' % (qname(hq['question']), ROLE.get(hq['role'], hq['role']),
                                 ' · Pflicht' if req == 'ALWAYS' else (' · bedingt Pflicht' if req == 'CONDITIONAL' else '')))
    srcs = '; '.join((s['document'] + (' ' + s['section'] if s.get('section') else '')) for s in h.get('sources', []))
    rows.append([h['code'], h['title'], h.get('category', ''), h.get('ui_group', ''),
                 h.get('aggregation_type', 'NONE'), h.get('hazard_factor', ''),
                 ', '.join(h.get('person_groups', [])), '\n'.join(qs), srcs,
                 ', '.join(h.get('review_ids', [])), len(rules_by.get(h['code'], []))])
write_rows(ws, rows)
ws.auto_filter.ref = 'A1:K%d' % (len(rows) + 1)

# ---- Blatt 4: Fragen -------------------------------------------------------
ws = wb.create_sheet('Fragen')
cols = ['Code', 'Nr.', 'Erhebungsbereich', 'Frage', 'Typ', 'Antwortoptionen', 'Hilfetext',
        'Sichtbar wenn', 'Gefährdungen']
header(ws, cols, [26, 7, 30, 60, 9, 44, 40, 40, 30])
hz_by_q = {}
for h in seed['hazards']:
    for hq in h.get('questions', []):
        hz_by_q.setdefault(hq['question'], []).append(h['code'])
rows = []
for q in seed['questions']:
    opts = '\n'.join(o['label'] for o in q.get('options', []))
    rows.append([q['code'], q.get('ui_number', ''), q.get('category', ''), q['text'], q['type'],
                 opts, q.get('help_text', ''), expr(q['visible_when']) if q.get('visible_when') else '',
                 ', '.join(sorted(set(hz_by_q.get(q['code'], []))))])
write_rows(ws, rows)
ws.auto_filter.ref = 'A1:I%d' % (len(rows) + 1)

# ---- Blatt 5: Lesehinweise -------------------------------------------------
ws = wb.create_sheet('Lesehinweise')
ws.column_dimensions['A'].width = 120
lines = [
    ('Cyber-GBU komponentenbasiert (CY) – Klärungsliste zur fachlichen Gegenlesung', True),
    ('Stand 02.09.2026 · Regelwerk %s · %d Fragen · %d Gefährdungen · %d Regeln · %d Klärungspunkte'
     % (seed['rule_version'], len(seed['questions']), len(seed['hazards']), len(seed['rules']), len(kl)), False),
    ('', False),
    ('Aufbau: Frage → Gefährdung (Rolle) → Regel → Stufe. Eine Gefährdung hat mehrere Fragen. Anlagenmerkmale '
     '(Blatt Fragen, Bereich A) filtern Gefährdungen auf „Nicht zutreffend". Fehlt eine Pflichtfrage, ist die '
     'Gefährdung „Unvollständig" – nie „Kein Risiko".', False),
    ('Quellen: Komponenten, Schnittstellen-Taxonomie, Maßnahmenkategorien und die 14 ZÜS-Prüfpunkte stammen aus '
     'dem zues-Katalog der App (EK-ZÜS B-002 rev. 5 Anhang 2, BA-017, DEKRA-Ausfüllhilfe 09/2024) und aus TRBS 1115 '
     'Teil 1; das Muster „vorhanden → Schnittstellen → Zugang frei → Maßnahmen" mit geteilten Zugangsfragen folgt '
     'als Methode der Schindler-Analyse (Kategorie 12, MC1/MC4). Anders als dort: sechs Zustände, ausdrückliche '
     'Kein-Risiko-Regel, Normbezug je Gefährdung, Stufe Hoch erreichbar. Alle Texte sind eigenständig formuliert.', False),
    ('Evidenz je Regel: HIGH_CONFIDENCE = Stufe und Maßnahme aus einer bestehenden App-Option übernommen; '
     'INFERRED = aus App-Katalog oder Schindler-Struktur abgeleitet, Text neu; HYPOTHESIS = eigener Vorschlag '
     'ohne Vorlage – bitte besonders prüfen.', False),
    ('Was zu tun ist: (1) Blatt Klaerungen – jede Zeile entscheiden (gelbe Spalten). (2) Blatt Regeln – '
     'stichprobenartig oder vollständig prüfen, Spalte „Prüfung OK?" und ggf. Korrektur eintragen. '
     '(3) Rückgabe der Datei; die Änderungen werden in mf_content/*.py eingepflegt und das Seed neu erzeugt.', False),
    ('Dieselben Inhalte sind im Prototyp „GBU-Bewertung" (Typ „GBU mehrfragig") interaktiv durchklickbar; '
     'dort zeigt jede Gefährdung ihre Fragen, die zutreffende Regel, Maßnahmen, Quellen und offene Klärungen.', False),
]
for i, (t, b) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=t)
    c.font = f(bold=b, size=11 if b else 10)
    c.alignment = WRAP

wb.save(OUT)
print('geschrieben:', OUT, '| Klärungen', len(kl), '| Regeln', len(seed['rules']))
