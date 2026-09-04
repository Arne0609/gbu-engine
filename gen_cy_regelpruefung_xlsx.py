# -*- coding: utf-8 -*-
"""Regelprüfung des Cyber-Typs: legt alle Regeln mit quality_status
REVIEW_REQUIRED (ohne Klärungspunkt entstandene Eigenregeln) zur fachlichen
Freigabe vor:  python3 gen_cy_regelpruefung_xlsx.py
Eingabe: norm_cyber_mf.json  ->  GBU_Cyber_Regelpruefung.xlsx

Blätter: Lesehinweise · Muster (6 Komponenten-Muster = 60 Regeln) ·
Regeln (alle REVIEW_REQUIRED-Regeln einzeln, Entscheidung als Auswahl).
Die Rückgabe wird mit apply_cy_regelpruefung.py in
cy_content/regelfreigabe.py übernommen (quality_status -> VERIFIED).
"""
import json, os, re
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
seed = json.load(open(os.path.join(HERE, 'norm_cyber_mf.json'), encoding='utf-8'))
OUT = os.path.join(HERE, 'GBU_Cyber_Regelpruefung.xlsx')

Q = {q['code']: q for q in seed['questions']}
H = {h['code']: h for h in seed['hazards']}
M = {m['code']: m for m in seed['measures']}
LAB = {'HIGH': 'Hoch', 'MEDIUM': 'Mittel', 'LOW': 'Niedrig', 'NO_RISK': 'Kein Risiko'}
EVID = {'HIGH_CONFIDENCE': 'hohe Sicherheit', 'INFERRED': 'abgeleitet', 'HYPOTHESIS': 'Hypothese'}

FONT = 'Arial'
def f(bold=False, color=None, size=10):
    return Font(name=FONT, bold=bold, color=color, size=size)
HEAD_FILL = PatternFill('solid', fgColor='1F3A5F')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')
STRIPE = PatternFill('solid', fgColor='F3F5F8')
PATTERN_FILL = PatternFill('solid', fgColor='E8F0E8')
thin = Side(style='thin', color='C9D1DB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top')


def val_label(qc, v):
    q = Q.get(qc)
    if q and q.get('type') == 'SELECT':
        for o in q.get('options', []):
            if o['value'] == v:
                return '„%s"' % o['label']
    if v is True: return 'Ja'
    if v is False: return 'Nein'
    return str(v)


def qname(qc):
    q = Q.get(qc)
    if not q: return qc
    return '%s %s' % (q.get('ui_number', ''), q['text'])


def expr(e):
    if 'all' in e: return '(' + ' UND '.join(expr(x) for x in e['all']) + ')'
    if 'any' in e: return '(' + ' ODER '.join(expr(x) for x in e['any']) + ')'
    if 'not' in e: return 'NICHT ' + expr(e['not'])
    op = e['operator']
    if op in ('ANSWERED', 'NOT_ANSWERED'):
        return qname(e['question']) + (' beantwortet' if op == 'ANSWERED' else ' unbeantwortet')
    v = e.get('value')
    if op in ('IN', 'NOT_IN'):
        return '%s %s [%s]' % (qname(e['question']), 'in' if op == 'IN' else 'nicht in',
                               ', '.join(val_label(e['question'], x) for x in v))
    sym = {'EQ': '=', 'NEQ': '≠', 'GT': '>', 'GTE': '≥', 'LT': '<', 'LTE': '≤'}[op]
    return '%s %s %s' % (qname(e['question']), sym, val_label(e['question'], v))


def header(ws, cols, widths):
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30


def write_rows(ws, rows, input_cols=(), pattern_col=None):
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = f()
            cell.alignment = WRAP
            cell.border = BORDER
            if c in input_cols:
                cell.fill = INPUT_FILL
            elif pattern_col and row[pattern_col - 1] and c == pattern_col:
                cell.fill = PATTERN_FILL
            elif r % 2 == 0:
                cell.fill = STRIPE


def measures_of(r):
    sofort = mittel = ''
    for mb in r.get('measures', []):
        m = M.get(mb['measure'])
        if not m: continue
        if mb.get('group_id') == 'sofort': sofort = m['title']
        elif mb.get('group_id') == 'mittel': mittel = m['title']
    return sofort, mittel


def sources_of(h):
    return '; '.join((s['document'] + (' ' + s['section'] if s.get('section') else ''))
                     for s in h.get('sources', []))


review = [r for r in seed['rules'] if r.get('quality_status') == 'REVIEW_REQUIRED']

# ---- Komponenten-Muster erkennen -------------------------------------------
# Die Komponenten CY-C01..C10 entstehen aus der Fabrik komponente(); ihre
# Regeln für lokale Schnittstellen (kabelgebunden/benutzer) sind bis auf den
# Fragen-Präfix identisch. Muster-Schlüssel: Bedingung mit qc_<slug>_ -> qc_X_.
def pattern_key(r):
    cond = re.sub(r'qc_[a-z]+_', 'qc_X_', json.dumps(r['condition'], sort_keys=True))
    return cond + '|' + r['result']

patterns = OrderedDict()
for r in review:
    if r['hazard'].startswith('CY-C') and int(r['hazard'][4:]) <= 10:
        patterns.setdefault(pattern_key(r), []).append(r)
# Muster-Nummern in Regel-Reihenfolge des ersten Auftretens
pattern_id = {k: 'M%d' % (i + 1) for i, k in enumerate(patterns)}
rule_pattern = {}
for k, rs in patterns.items():
    for r in rs:
        rule_pattern[r['code']] = pattern_id[k]

wb = Workbook()

# ---- Blatt 1: Lesehinweise -------------------------------------------------
ws = wb.active
ws.title = 'Lesehinweise'
ws.column_dimensions['A'].width = 120
n_pat_rules = sum(len(v) for v in patterns.values())
lines = [
    ('Cyber-GBU komponentenbasiert (CY) – Regelprüfung zur Freigabe', True),
    ('Stand 03.09.2026 · Regelwerk %s · %d Regeln gesamt, davon %d VERIFIED (über entschiedene '
     'Klärungen K-C01…K-C25) und %d REVIEW_REQUIRED (diese Liste).'
     % (seed['rule_version'], len(seed['rules']),
        len(seed['rules']) - len(review), len(review)), False),
    ('', False),
    ('Was hier vorliegt: Regeln, an denen KEIN Klärungspunkt hing – also die Eigenregeln, deren '
     'Ableitung beim Entwurf als unstrittig galt (Evidenz „hohe Sicherheit" = direkt aus TRBS 1115-1 / '
     'B-002 / BA-017 ableitbar; „abgeleitet" = Stufe nach Nohl-Logik der anderen GBU-Typen gesetzt). '
     'Sie sind fachlich noch nicht freigegeben und laufen deshalb mit quality_status REVIEW_REQUIRED.', False),
    ('', False),
    ('Blatt „Muster": %d der %d Regeln sind die %d Komponenten-Muster der Fabrik komponente() für '
     'lokale Schnittstellen (kabelgebunden / Benutzerschnittstelle) bzw. die Auffangregel „Kein Risiko" – '
     'je Muster einmal für alle 10 Komponenten CY-C01…C10 entscheiden. Eine Entscheidung im Blatt '
     '„Muster" gilt für alle zugehörigen Regeln; im Blatt „Regeln" muss dann nichts mehr eingetragen '
     'werden (Abweichungen für einzelne Komponenten dort in „Korrektur" notieren).'
     % (n_pat_rules, len(review), len(patterns)), False),
    ('Blatt „Regeln": alle %d Regeln einzeln (Filter über Spalte „Muster": leer = Einzelregel). '
     'Die %d Einzelregeln (Zugang, Netz, Kanäle CY-C11…C14, Organisation) bitte je Zeile entscheiden.'
     % (len(review), len(review) - n_pat_rules), False),
    ('', False),
    ('Entscheidung (gelbe Spalte, Auswahl): „Freigeben" = Regel stimmt, wird VERIFIED · '
     '„Ändern" = Stufe/Bedingung/Maßnahme anpassen, Festlegung in „Korrektur" · '
     '„Streichen" = Regel entfällt (Achtung: nur sinnvoll, wenn eine andere Regel den Fall abdeckt – '
     'die Engine bleibt fail-closed, sonst ergibt der Fall „Unvollständig/Regellücke").', False),
    ('', False),
    ('Lesehilfe Bedingung: Fragen mit Nr. und Text; „in [A, B]" = Antwort ist eine der genannten; '
     '„Zugang frei" = eine der drei Zugangsfragen (Steuerung / Triebwerksraum / Schacht frei zugänglich) '
     'mit Ja. Prio: höhere Zahl gewinnt (spezifischere Regel zuerst); Prio 1 ist die Auffangregel '
     '„Kein Risiko", die nur greift, wenn keine Mangelregel zutrifft und alle Pflichtfragen beantwortet sind.', False),
    ('Stufenlogik Komponente (zur Einordnung): Fernzugriff/kabellos ohne Maßnahmen = Hoch (K-C06, '
     'VERIFIED) · lokal ohne Maßnahmen = Mittel · teilweise = Niedrig, mit freiem Zugang Mittel · '
     'umgesetzt = Kein Risiko, mit freiem Zugang Niedrig · unabhängige Sicherheitseinrichtung deckelt '
     'auf Mittel/Niedrig (K-C07…K-C09, VERIFIED).', False),
]
for i, (t, b) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=t)
    c.font = f(bold=b, size=11 if b else 10)
    c.alignment = WRAP

# ---- Blatt 2: Muster -------------------------------------------------------
ws = wb.create_sheet('Muster')
cols = ['Muster', 'Regeln', 'Komponenten', 'Prio', 'Bedingung (X = Komponente)', 'Stufe', 'Evidenz',
        'Sofortmaßnahme', 'Mittelfristige Maßnahme', 'Begründung im Entwurf', 'Entscheidung', 'Korrektur']
header(ws, cols, [8, 8, 30, 6, 70, 12, 14, 40, 40, 44, 14, 40])
GRUND = {
    'MEDIUM|keine': 'Lokale Schnittstelle ohne Maßnahmen: Angriff setzt physische Nähe voraus, '
                    'daher eine Stufe unter Fernzugriff (Hoch). TRBS 1115-1 4.5.2 verlangt Maßnahmen '
                    'für jede erreichbare Schnittstelle.',
    'MEDIUM|teilweise': 'Maßnahmen nur teilweise UND Zugang frei: die Schnittstelle ist ohne Hürde '
                        'erreichbar, Restlücke wird wie „keine Maßnahmen" behandelt.',
    'LOW|teilweise': 'Maßnahmen teilweise, Zugang gesichert: Restrisiko, Nachbesserung terminieren.',
    'LOW|umgesetzt': 'Maßnahmen umgesetzt, aber Zugang frei: die organisatorisch-technischen Maßnahmen '
                     'stehen, die physische Zugangskontrolle fehlt (siehe Gefährdungen Zugang).',
    'NO_RISK|umgesetzt': 'Maßnahmen umgesetzt, Zugang gesichert: Sollzustand nach TRBS 1115-1.',
    'NO_RISK|ANSWERED': 'Auffangregel: keine Mangelregel trifft zu, Pflichtfragen beantwortet.',
}
def grund_of(r):
    cs = json.dumps(r['condition'])
    if '"ANSWERED"' in cs: return GRUND['NO_RISK|ANSWERED']
    for m in ('keine', 'teilweise', 'umgesetzt'):
        if '"%s"' % m in cs:
            return GRUND.get('%s|%s' % (r['result'], m), '')
    return ''

rows = []
for k, rs in patterns.items():
    r0 = rs[0]
    sofort, mittel = measures_of(r0)
    # Anzeige mit dem ersten Vertreter (CY-C01 = Aufzugssteuerung), Komponenten-
    # name und Fragennummer neutralisiert: „Aufzugssteuerung" -> „X", 3.1.2 -> 3.X.2
    txt = expr(r0['condition'])
    for qc in re.findall(r'qc_[a-z]+_(?:schnittstelle|massnahmen)', json.dumps(r0['condition'])):
        q = Q[qc]
        name = re.search(r'„([^"]+)"', q['text'])
        if name:
            txt = txt.replace(name.group(1), 'X')
        txt = txt.replace(q.get('ui_number', ''), re.sub(r'^3\.\d+\.', '3.X.', q.get('ui_number', '')))
    rows.append([pattern_id[k], len(rs), ', '.join(r['hazard'] for r in rs), r0['priority'],
                 txt, LAB[r0['result']], EVID.get(r0.get('evidence'), ''),
                 sofort, mittel, grund_of(r0), None, None])
write_rows(ws, rows, input_cols=(11, 12))
dv = DataValidation(type='list', formula1='"Freigeben,Ändern,Streichen"', allow_blank=True)
ws.add_data_validation(dv)
dv.add('K2:K%d' % (len(rows) + 1))

# ---- Blatt 3: Regeln -------------------------------------------------------
ws = wb.create_sheet('Regeln')
cols = ['Nr.', 'Gefährdung', 'Titel', 'Erhebungsbereich', 'Regel', 'Prio', 'Bedingung', 'Stufe',
        'Evidenz', 'Muster', 'Sofortmaßnahme', 'Mittelfristige Maßnahme', 'Hinweis im Entwurf',
        'Norm-/Quellenbezug der Gefährdung', 'Entscheidung', 'Korrektur']
header(ws, cols, [5, 10, 36, 22, 13, 6, 70, 12, 14, 8, 40, 40, 34, 40, 14, 40])
rows = []
for i, r in enumerate(review, 1):
    h = H[r['hazard']]
    sofort, mittel = measures_of(r)
    rows.append([i, r['hazard'], h['title'], h.get('ui_group', ''), r['code'], r['priority'],
                 expr(r['condition']), LAB[r['result']], EVID.get(r.get('evidence'), ''),
                 rule_pattern.get(r['code'], ''), sofort, mittel, r.get('notes', ''),
                 sources_of(h), None, None])
write_rows(ws, rows, input_cols=(15, 16), pattern_col=10)
dv2 = DataValidation(type='list', formula1='"Freigeben,Ändern,Streichen"', allow_blank=True)
ws.add_data_validation(dv2)
dv2.add('O2:O%d' % (len(rows) + 1))
ws.auto_filter.ref = 'A1:P%d' % (len(rows) + 1)
n = len(rows) + 1
r0 = n + 2
ws.cell(row=r0, column=1, value='Stand').font = f(bold=True)
ws.cell(row=r0 + 1, column=1, value='Regeln').font = f()
ws.cell(row=r0 + 1, column=2, value='=COUNTA(E2:E%d)' % n).font = f()
ws.cell(row=r0 + 2, column=1, value='davon Muster').font = f()
ws.cell(row=r0 + 2, column=2, value='=COUNTIF(J2:J%d,"M*")' % n).font = f()
ws.cell(row=r0 + 3, column=1, value='Einzelregeln entschieden').font = f()
ws.cell(row=r0 + 3, column=2, value='=COUNTIFS(J2:J%d,"",O2:O%d,"<>")' % (n, n)).font = f()

wb.save(OUT)
print('geschrieben:', OUT, '| Regeln:', len(review), '| Muster:', len(patterns),
      '(%d Regeln) | Einzelregeln: %d' % (n_pat_rules, len(review) - n_pat_rules))
